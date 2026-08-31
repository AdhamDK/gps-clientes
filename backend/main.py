"""FastAPI app — Cliente CRUD + Search + Import/Export (PR3) with UUID spec.

Endpoints:
  POST   /clientes
  GET    /clientes?q=&zona=&page=&limit=
  GET    /clientes/{id}
  PATCH  /clientes/{id}
  DELETE /clientes/{id} (soft delete)
  POST   /clientes/import  (multipart xlsx openpyxl 20->11)
  GET    /clientes/export?formato=xlsx|pdf&columnas=  (streaming)
"""

import io
import logging
import re
import uuid
from datetime import date, datetime, timezone

import httpx
import openpyxl
import os

from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from . import models, schemas, vroom_client
from .database import Base, engine, get_db, normalize_nfd, generate_uuid, utc_now_iso

logger = logging.getLogger(__name__)

RIF_REGEX = re.compile(r"^[JVEGP]\d{7,9}$")
EXPORT_WHITELIST = [
    "id",
    "nombre",
    "nombre_normalizado",
    "rif",
    "rif_cedula",
    "telefono",
    "direccion",
    "direccion_original",
    "empresa",
    "zona",
    "lat",
    "lng",
    "latitud",
    "longitud",
    "texto_breve",
    "is_flagged",
    "has_gps_fix",
    "updated_at",
    "sync_status",
    "deleted",
]
EXPORT_WHITELIST_SET = set(EXPORT_WHITELIST)

MAX_IMPORT_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
EXPORT_MAX_ROWS = 5000

# Ensure tables exist (greenfield Base.metadata.create_all) + migrate legacy columns
Base.metadata.create_all(bind=engine)
try:
    from sqlalchemy import inspect as _inspect, text as _text
    from sqlalchemy.schema import CreateIndex as _CreateIndex, CreateTable as _CreateTable

    _insp = _inspect(engine)

    # --- Legacy INTEGER PK migration: clientes.id INTEGER -> VARCHAR (TEXT UUID) ---
    try:
        if _insp.has_table("clientes"):
            _cols_map = {c["name"]: c for c in _insp.get_columns("clientes")}
            _id_type_str = str(_cols_map.get("id", {}).get("type", "")) if "id" in _cols_map else ""
            if _id_type_str.upper() == "INTEGER":
                with engine.begin() as _conn:
                    for _idx in ["ix_clientes_nombre_normalizado", "ix_clientes_rif", "ix_clientes_rif_cedula", "ix_clientes_has_gps_fix", "ix_clientes_sync_status", "ix_clientes_deleted"]:
                        try:
                            _conn.execute(_text(f"DROP INDEX IF EXISTS {_idx}"))
                        except Exception as e:
                            logger.warning("migration drop clientes index failed: %s", e, exc_info=True)
                    _conn.execute(_text("ALTER TABLE clientes RENAME TO clientes_legacy_backup"))
                    models.Cliente.__table__.create(bind=_conn)
                    _legacy_cols = {c["name"] for c in _inspect(_conn).get_columns("clientes_legacy_backup")}
                    _new_cols = [c.name for c in models.Cliente.__table__.columns]
                    _sel = []
                    for _c in _new_cols:
                        if _c == "id":
                            _sel.append("CAST(id AS TEXT) as id")
                        elif _c in _legacy_cols:
                            _sel.append(f'"{_c}"')
                        else:
                            if _c in ("sync_status", "deleted", "is_flagged", "has_gps_fix"):
                                _sel.append(f"0 as \"{_c}\"")
                            elif _c == "updatedAt":
                                _sel.append("NULL as \"updatedAt\"")
                            else:
                                _sel.append(f"NULL as \"{_c}\"")
                    _sql = f'INSERT INTO clientes ("' + '","'.join(_new_cols) + '") SELECT ' + ", ".join(_sel) + ' FROM clientes_legacy_backup'
                    _conn.execute(_text(_sql))
                    for _idx in models.Cliente.__table__.indexes:
                        try:
                            _conn.execute(_text(str(_CreateIndex(_idx).compile(dialect=engine.dialect))))
                        except Exception as e:
                            logger.warning("migration create clientes index failed: %s", e, exc_info=True)
                    _conn.execute(_text("DROP TABLE clientes_legacy_backup"))
                _insp = _inspect(engine)
    except Exception as e:
        logger.warning("clientes legacy INTEGER migration failed: %s", e, exc_info=True)

    # --- Legacy rutas_hoy.cliente_id INTEGER -> VARCHAR ---
    try:
        if _insp.has_table("rutas_hoy"):
            _cols_map = {c["name"]: c for c in _insp.get_columns("rutas_hoy")}
            _ct = str(_cols_map.get("cliente_id", {}).get("type", "")) if "cliente_id" in _cols_map else ""
            if _ct.upper() == "INTEGER":
                with engine.begin() as _conn:
                    for _idx in ["ix_rutas_hoy_fecha_orden", "ix_rutas_hoy_cliente_id", "ix_rutas_hoy_fecha_entregado"]:
                        try:
                            _conn.execute(_text(f"DROP INDEX IF EXISTS {_idx}"))
                        except Exception as e:
                            logger.warning("migration drop rutas_hoy index failed: %s", e, exc_info=True)
                    _conn.execute(_text("ALTER TABLE rutas_hoy RENAME TO rutas_hoy_legacy_backup"))
                    models.RutasHoy.__table__.create(bind=_conn)
                    _legacy_cols = {c["name"] for c in _inspect(_conn).get_columns("rutas_hoy_legacy_backup")}
                    _new_cols = [c.name for c in models.RutasHoy.__table__.columns]
                    _sel = []
                    for _c in _new_cols:
                        if _c == "cliente_id":
                            _sel.append("CAST(cliente_id AS TEXT) as cliente_id")
                        elif _c in _legacy_cols:
                            _sel.append(f'"{_c}"')
                        else:
                            _sel.append(f"NULL as \"{_c}\"")
                    _sql = f'INSERT INTO rutas_hoy ("' + '","'.join(_new_cols) + '") SELECT ' + ", ".join(_sel) + ' FROM rutas_hoy_legacy_backup'
                    _conn.execute(_text(_sql))
                    for _idx in models.RutasHoy.__table__.indexes:
                        try:
                            _conn.execute(_text(str(_CreateIndex(_idx).compile(dialect=engine.dialect))))
                        except Exception as e:
                            logger.warning("migration create rutas_hoy index failed: %s", e, exc_info=True)
                    _conn.execute(_text("DROP TABLE rutas_hoy_legacy_backup"))
                _insp = _inspect(engine)
    except Exception as e:
        logger.warning("rutas_hoy legacy INTEGER migration failed: %s", e, exc_info=True)

    def _migrate_table(table_name, table):
        if not _insp.has_table(table_name):
            return
        existing = {c["name"] for c in _insp.get_columns(table_name)}
        # collect missing columns vs model definition
        missing = [col for col in table.columns if col.name not in existing]
        if not missing:
            return
        with engine.begin() as _conn:
            for col in missing:
                coltype = col.type.compile(dialect=engine.dialect)
                ddl_type = str(coltype)
                # SQLite-friendly type overrides for key columns
                if col.name == "updated_at":
                    ddl_type = "VARCHAR"
                elif col.name == "updatedAt":
                    ddl_type = "DATETIME"
                elif col.name in ("sync_status", "deleted"):
                    ddl_type = "INTEGER"
                elif col.name in ("latitud", "longitud", "lat", "lng"):
                    ddl_type = "REAL"
                elif col.name in ("is_flagged", "has_gps_fix"):
                    ddl_type = "BOOLEAN"
                # Build DEFAULT clause for non-callable defaults
                default_clause = ""
                try:
                    if col.default is not None and hasattr(col.default, "arg"):
                        arg = col.default.arg
                        if not callable(arg):
                            if isinstance(arg, bool):
                                default_clause = f" DEFAULT {1 if arg else 0}"
                            elif isinstance(arg, int):
                                default_clause = f" DEFAULT {arg}"
                            elif isinstance(arg, str):
                                default_clause = f" DEFAULT '{arg}'"
                    if col.server_default is not None:
                        sd = str(col.server_default.arg)
                        if "DEFAULT" not in sd.upper():
                            default_clause += f" DEFAULT {sd}"
                        else:
                            default_clause += f" {sd}"
                except Exception as e:
                    logger.warning("migrate default clause failed: %s", e, exc_info=True)
                    default_clause = ""
                # Ensure safe NOT NULL handling: avoid NOT NULL without default on existing rows
                not_null = ""
                if not col.nullable and default_clause.strip() != "":
                    not_null = " NOT NULL"
                # Force defaults for sync_status/deleted to allow NOT NULL
                if col.name in ("sync_status", "deleted"):
                    if not default_clause.strip():
                        default_clause = " DEFAULT 0"
                    not_null = " NOT NULL"
                # updated_at/updatedAt stay nullable for backfill compat
                if col.name in ("updated_at", "updatedAt"):
                    not_null = ""
                ddl = f'ALTER TABLE "{table_name}" ADD COLUMN "{col.name}" {ddl_type}{not_null}{default_clause}'
                _conn.execute(_text(ddl))

    _migrate_table("clientes", models.Cliente.__table__)
    _migrate_table("rutas_hoy", models.RutasHoy.__table__)
except Exception as e:
    logger.warning("table migration failed: %s", e, exc_info=True)

app = FastAPI(title="GPS Clientes API", version="0.3.0")

# Security: explicit CORS whitelist — not "*" — to block file:// and arbitrary origins in WebView
# Includes legacy dev origins (3000/5173) for test compatibility; still no wildcard.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:8000","http://127.0.0.1:8000","http://localhost:3000","http://localhost:5173","https://appassets.androidplatform.net","http://10.0.2.2:8000"],
    allow_origin_regex=None,
    allow_credentials=False,
    allow_methods=["GET","POST","PATCH","DELETE","OPTIONS"],
    allow_headers=["Content-Type","Authorization"],
)


def _to_flagged(empresa: str | None) -> bool:
    return bool(empresa and "#" in empresa)


def _to_has_gps_fix(lat: float | None, lng: float | None) -> bool:
    return lat is not None and lng is not None


def _resolve_lat_lng(payload_lat, payload_lng, payload_latitud, payload_longitud):
    """Resolve spec latitud/longitud vs legacy lat/lng — coalesce, prefer spec."""
    lat = payload_latitud if payload_latitud is not None else payload_lat
    lng = payload_longitud if payload_longitud is not None else payload_lng
    # ensure both pairs synced
    return lat, lng


def _resolve_rif(rif, rif_cedula):
    """Coalesce rif and rif_cedula — keep both in sync."""
    val = rif_cedula if rif_cedula is not None else rif
    return val, val


@app.post("/clientes", response_model=schemas.ClienteRead, status_code=201)
def create_cliente(payload: schemas.ClienteCreate, db: Session = Depends(get_db)):
    nombre_normalizado = normalize_nfd(payload.nombre)
    rif_val, rif_cedula_val = _resolve_rif(payload.rif, getattr(payload, "rif_cedula", None))
    lat_val, lng_val = _resolve_lat_lng(payload.lat, payload.lng, getattr(payload, "latitud", None), getattr(payload, "longitud", None))
    # has_gps_fix derived
    has_fix = _to_has_gps_fix(lat_val, lng_val)
    # id generation
    cliente_id = getattr(payload, "id", None)
    if not cliente_id:
        cliente_id = generate_uuid()
    # ensure lowercase UUID
    try:
        cliente_id = str(uuid.UUID(str(cliente_id))).lower()
    except Exception as e:
        logger.warning("invalid cliente_id fallback to generate_uuid: %s", e, exc_info=True)
        cliente_id = generate_uuid()
    updated = getattr(payload, "updated_at", None)
    if not updated:
        updated = datetime.now(timezone.utc)
    elif isinstance(updated, str):
        try:
            updated = datetime.fromisoformat(updated.replace("Z", "+00:00"))
        except Exception as e:
            logger.warning("invalid updated_at parse, using now: %s", e, exc_info=True)
            updated = datetime.now(timezone.utc)
    sync_st = getattr(payload, "sync_status", None)
    if sync_st is None:
        sync_st = 0
    del_flag = getattr(payload, "deleted", None)
    if del_flag is None:
        del_flag = 0
    cliente = models.Cliente(
        id=cliente_id,
        nombre=payload.nombre.strip(),
        nombre_normalizado=nombre_normalizado,
        rif=rif_val,
        rif_cedula=rif_cedula_val,
        telefono=payload.telefono,
        direccion=payload.direccion,
        direccion_original=getattr(payload, "direccion_original", None) if getattr(payload, "direccion_original", None) is not None else payload.direccion,
        empresa=payload.empresa,
        zona=payload.zona,
        lat=lat_val,
        lng=lng_val,
        latitud=lat_val,
        longitud=lng_val,
        texto_breve=payload.texto_breve,
        is_flagged=_to_flagged(payload.empresa),
        has_gps_fix=has_fix,
        updated_at=updated,
        sync_status=sync_st,
        deleted=del_flag,
    )
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


class ClienteListResponse(schemas.ClienteRead):
    pass


@app.get("/clientes")
def list_clientes(
    q: str | None = Query(default=None, description="NFD search on nombre/texto_breve"),
    zona: str | None = Query(default=None, description="Exact zona match"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=20, ge=1, le=500),
    include_deleted: bool = Query(default=False, description="include soft-deleted"),
    db: Session = Depends(get_db),
):
    base_filter_deleted = [] if include_deleted else [models.Cliente.deleted == 0]
    if q is not None and q.strip() != "":
        q_norm = normalize_nfd(q)
        escaped = q_norm.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        like_pattern = f"%{escaped}%"
        query = db.query(models.Cliente).filter(*base_filter_deleted)
        if zona is not None:
            query = query.filter(models.Cliente.zona == zona)
        query = query.filter(
            or_(
                models.Cliente.nombre_normalizado.like(like_pattern, escape="\\"),
                models.Cliente.texto_breve.like(like_pattern, escape="\\"),
            )
        )
        total = query.count()
        offset = (page - 1) * limit
        items = query.order_by(models.Cliente.id).offset(offset).limit(limit).all()
        serialized = [schemas.ClienteRead.model_validate(c).model_dump(mode="json") for c in items]
        return {"items": serialized, "total": total, "page": page, "limit": limit}

    query = db.query(models.Cliente).filter(*base_filter_deleted)
    if zona is not None:
        query = query.filter(models.Cliente.zona == zona)

    total = query.count()
    offset = (page - 1) * limit
    items = query.order_by(models.Cliente.id).offset(offset).limit(limit).all()
    serialized = [schemas.ClienteRead.model_validate(c).model_dump(mode="json") for c in items]
    return {"items": serialized, "total": total, "page": page, "limit": limit}


def _export_value(cliente: models.Cliente, col: str) -> str:
    if col == "id":
        return str(cliente.id)
    if col == "nombre":
        return cliente.nombre or ""
    if col == "nombre_normalizado":
        return cliente.nombre_normalizado or ""
    if col == "rif":
        return cliente.rif or cliente.rif_cedula or ""
    if col == "rif_cedula":
        return cliente.rif_cedula or cliente.rif or ""
    if col == "telefono":
        return cliente.telefono or ""
    if col == "direccion":
        return cliente.direccion or ""
    if col == "direccion_original":
        return cliente.direccion_original or ""
    if col == "empresa":
        return cliente.empresa or ""
    if col == "zona":
        return cliente.zona or ""
    if col == "lat":
        v = cliente.lat if cliente.lat is not None else cliente.latitud
        return str(v) if v is not None else ""
    if col == "lng":
        v = cliente.lng if cliente.lng is not None else cliente.longitud
        return str(v) if v is not None else ""
    if col == "latitud":
        v = cliente.latitud if cliente.latitud is not None else cliente.lat
        return str(v) if v is not None else ""
    if col == "longitud":
        v = cliente.longitud if cliente.longitud is not None else cliente.lng
        return str(v) if v is not None else ""
    if col == "texto_breve":
        return cliente.texto_breve or ""
    if col == "is_flagged":
        return str(cliente.is_flagged)
    if col == "has_gps_fix":
        return str(cliente.has_gps_fix)
    if col == "updated_at":
        if cliente.updated_at:
            return str(cliente.updated_at) if isinstance(cliente.updated_at, str) else cliente.updated_at.isoformat()
        return ""
    if col == "sync_status":
        return str(cliente.sync_status) if cliente.sync_status is not None else "0"
    if col == "deleted":
        return str(cliente.deleted) if cliente.deleted is not None else "0"
    return ""


@app.get("/clientes/export")
def export_clientes(
    formato: str = Query(default="xlsx", pattern="^(xlsx|pdf)$", description="xlsx or pdf"),
    columnas: str | None = Query(default=None, description="comma-separated whitelist"),
    db: Session = Depends(get_db),
):
    if columnas is not None and columnas.strip() != "":
        picked = [c.strip() for c in columnas.split(",") if c.strip() != ""]
        invalid = [c for c in picked if c not in EXPORT_WHITELIST_SET]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid columns: {invalid}. Whitelist: {EXPORT_WHITELIST}")
        cols = picked
    else:
        cols = EXPORT_WHITELIST
    count = db.query(models.Cliente).filter(models.Cliente.deleted == 0).count()
    if count > EXPORT_MAX_ROWS:
        raise HTTPException(status_code=413, detail="Demasiados registros — filtra por zona/q")
    items = db.query(models.Cliente).filter(models.Cliente.deleted == 0).order_by(models.Cliente.id).all()
    if formato == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        for idx, col in enumerate(cols, start=1):
            ws.cell(row=1, column=idx, value=col)
            ws.cell(row=1, column=idx).font = openpyxl.styles.Font(bold=True)
        for r_idx, cliente in enumerate(items, start=2):
            for c_idx, col in enumerate(cols, start=1):
                ws.cell(row=r_idx, column=c_idx, value=_export_value(cliente, col))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="clientes.xlsx"'}
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail=f"reportlab not installed: {exc}") from exc
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 36
    row_h = 18
    header_h = 22
    col_w = (width - 2 * margin) / max(len(cols), 1)
    col_w = max(col_w, 60)
    y = height - margin

    def draw_header(y_pos: float) -> float:
        c.setFont("Helvetica-Bold", 7)
        for idx, col in enumerate(cols):
            x = margin + idx * col_w
            label = col[: int(col_w / 5)] if len(col) * 4 > col_w else col
            c.drawString(x + 2, y_pos - 12, label)
            c.rect(x, y_pos - header_h, col_w, header_h, stroke=1, fill=0)
        return y_pos - header_h

    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y, "Clientes Export")
    y -= 20
    y = draw_header(y)
    c.setFont("Helvetica", 6)
    for cliente in items:
        if y - row_h < margin:
            c.showPage()
            y = height - margin
            y = draw_header(y)
            c.setFont("Helvetica", 6)
        for idx, col in enumerate(cols):
            x = margin + idx * col_w
            val = _export_value(cliente, col)
            max_chars = max(int(col_w / 4), 8)
            txt = val[:max_chars] + ("..." if len(val) > max_chars else "")
            c.drawString(x + 2, y - 12, txt)
            c.rect(x, y - row_h, col_w, row_h, stroke=1, fill=0)
        y -= row_h
    c.showPage()
    c.save()
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="clientes.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


@app.get("/clientes/{cliente_id}", response_model=schemas.ClienteRead)
def get_cliente(cliente_id: str, db: Session = Depends(get_db)):
    cliente = db.get(models.Cliente, cliente_id)
    if not cliente or cliente.deleted == 1:
        raise HTTPException(status_code=404, detail="Cliente not found")
    return cliente


@app.patch("/clientes/{cliente_id}", response_model=schemas.ClienteRead)
def update_cliente(cliente_id: str, payload: schemas.ClienteUpdate, db: Session = Depends(get_db)):
    cliente = db.get(models.Cliente, cliente_id)
    if not cliente or cliente.deleted == 1:
        raise HTTPException(status_code=404, detail="Cliente not found")

    data = payload.model_dump(exclude_unset=True)

    if "nombre" in data and data["nombre"] is not None:
        cliente.nombre = data["nombre"].strip()
        cliente.nombre_normalizado = normalize_nfd(data["nombre"])

    if "rif" in data:
        cliente.rif = data["rif"]
        cliente.rif_cedula = data["rif"]
    if "rif_cedula" in data:
        # rif_cedula takes precedence if both present
        cliente.rif_cedula = data["rif_cedula"]
        cliente.rif = data["rif_cedula"]
    if "telefono" in data:
        cliente.telefono = data["telefono"]
    if "direccion" in data:
        cliente.direccion = data["direccion"]
    if "empresa" in data:
        cliente.empresa = data["empresa"]
        cliente.is_flagged = _to_flagged(data["empresa"])
    if "zona" in data:
        cliente.zona = data["zona"]
    # handle lat/lng / latitud/longitud coalescence
    lat_updated = False
    lng_updated = False
    if "lat" in data:
        cliente.lat = data["lat"]
        cliente.latitud = data["lat"]
        lat_updated = True
    if "lng" in data:
        cliente.lng = data["lng"]
        cliente.longitud = data["lng"]
        lng_updated = True
    if "latitud" in data:
        cliente.latitud = data["latitud"]
        cliente.lat = data["latitud"]
        lat_updated = True
    if "longitud" in data:
        cliente.longitud = data["longitud"]
        cliente.lng = data["longitud"]
        lng_updated = True
    if lat_updated or lng_updated:
        # prefer lat/lng after sync
        final_lat = cliente.lat if cliente.lat is not None else cliente.latitud
        final_lng = cliente.lng if cliente.lng is not None else cliente.longitud
        cliente.has_gps_fix = _to_has_gps_fix(final_lat, final_lng)
    if "texto_breve" in data:
        cliente.texto_breve = data["texto_breve"]
    if "sync_status" in data and data["sync_status"] is not None:
        cliente.sync_status = data["sync_status"]
    else:
        # any update marks pending unless explicitly set
        cliente.sync_status = 0
    if "deleted" in data and data["deleted"] is not None:
        cliente.deleted = data["deleted"]
    # always bump updated_at unless explicit
    if "updated_at" in data and data["updated_at"] is not None:
        val = data["updated_at"]
        if isinstance(val, str):
            try:
                val = datetime.fromisoformat(val.replace("Z", "+00:00"))
            except Exception as e:
                logger.warning("invalid updated_at parse: %s", e, exc_info=True)
                val = datetime.now(timezone.utc)
        cliente.updated_at = val
    else:
        cliente.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(cliente)
    return cliente


@app.delete("/clientes/{cliente_id}", status_code=204)
def delete_cliente(cliente_id: str, db: Session = Depends(get_db)):
    cliente = db.get(models.Cliente, cliente_id)
    if not cliente or cliente.deleted == 1:
        raise HTTPException(status_code=404, detail="Cliente not found")
    # soft delete per spec
    cliente.deleted = 1
    cliente.sync_status = 0
    cliente.updated_at = datetime.now(timezone.utc)
    db.commit()
    return None


# ---------------------------------------------------------------------------
# Health + Route Optimization (PR4 VROOM + OSRM)
# ---------------------------------------------------------------------------


@app.get("/health")
def health():
    """Liveness probe — always 200, checks VROOM :3000 and OSRM :5000 inline."""
    vroom = vroom_client.check_vroom()
    osrm = vroom_client.check_osrm()
    return {"status": "ok", "vroom": vroom, "osrm": osrm}


class OptimizeRequest(BaseModel):
    cliente_ids: list[str]
    start: list[float] | None = None  # [lng, lat]


@app.post("/rutas/optimizar")
def optimizar_rutas(payload: OptimizeRequest, db: Session = Depends(get_db)):
    if not payload.cliente_ids:
        raise HTTPException(status_code=400, detail="cliente_ids required")
    clientes = db.query(models.Cliente).filter(models.Cliente.id.in_(payload.cliente_ids), models.Cliente.deleted == 0).all()
    by_id = {str(c.id): c for c in clientes}
    missing = [cid for cid in payload.cliente_ids if str(cid) not in by_id]
    if missing:
        raise HTTPException(status_code=404, detail=f"Clientes not found: {missing}")
    # also check has_gps_fix / lat
    no_gps = []
    for c in clientes:
        lat = c.lat if c.lat is not None else c.latitud
        lng = c.lng if c.lng is not None else c.longitud
        if lat is None or lng is None or not c.has_gps_fix:
            no_gps.append(str(c.id))
    if no_gps:
        raise HTTPException(status_code=400, detail=f"Clientes without GPS fix: {no_gps}")

    # build jobs using resolved coords
    jobs = []
    for c in clientes:
        lat = c.lat if c.lat is not None else c.latitud
        lng = c.lng if c.lng is not None else c.longitud
        jobs.append({"id": str(c.id), "location": [lng, lat]})
    loc_by_id = {j["id"]: j["location"] for j in jobs}

    vroom_data = vroom_client.optimize_via_vroom(jobs, start=payload.start)

    # Parse orden from steps type job
    orden: list[str] = []
    try:
        routes = vroom_data.get("routes") or []
        if routes:
            steps = routes[0].get("steps") or []
            for s in steps:
                if s.get("type") == "job" and "id" in s:
                    orden.append(str(s["id"]))
    except Exception as e:
        logger.warning("VROOM parse orden failed: %s", e, exc_info=True)
        orden = []

    if not orden:
        orden = [str(x) for x in payload.cliente_ids]

    distance = 0
    duration = 0
    summary = vroom_data.get("summary") or {}
    if summary:
        distance = summary.get("distance", 0)
        duration = summary.get("duration", 0)
    else:
        try:
            r0 = (vroom_data.get("routes") or [{}])[0]
            distance = r0.get("distance", 0) or r0.get("summary", {}).get("distance", 0)
            duration = r0.get("duration", 0) or r0.get("summary", {}).get("duration", 0)
        except Exception as e:
            logger.warning("VROOM parse distance/duration failed: %s", e, exc_info=True)

    coords = [loc_by_id[cid] for cid in orden if cid in loc_by_id]
    geometry_result: dict | None = None
    geometry = None
    try:
        geometry_result = vroom_client.fetch_geometry(coords)
        geometry = geometry_result.get("geometry")
        if (not distance or not duration) and geometry_result:
            distance = geometry_result.get("distance", distance)
            duration = geometry_result.get("duration", duration)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.warning("OSRM geometry failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=502, detail=f"OSRM geometry failed: {exc}") from exc

    # VROOM/OSRM already called outside transaction — atomic DB write
    today = date.today()
    if db.in_transaction():
        db.commit()
    with db.begin():
        db.query(models.RutasHoy).filter(models.RutasHoy.fecha == today).delete()
        for idx, cid in enumerate(orden):
            db.add(models.RutasHoy(cliente_id=str(cid), orden=idx, fecha=today, entregado=False, delivered_at=None))
    # commit automatic via db.begin()

    return {"orden": orden, "distance": distance, "duration": duration, "geometry": geometry}


@app.get("/rutas/hoy")
def get_rutas_hoy(fecha: str | None = Query(default=None, description="YYYY-MM-DD"), entregado: bool | None = Query(default=None, description="filter by entregado"), db: Session = Depends(get_db)):
    if fecha is None:
        target = date.today()
    else:
        try:
            target = date.fromisoformat(fecha)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid fecha format YYYY-MM-DD: {fecha}") from exc
    q = db.query(models.RutasHoy).options(joinedload(models.RutasHoy.cliente)).filter(models.RutasHoy.fecha == target)
    if entregado is not None:
        q = q.filter(models.RutasHoy.entregado == entregado)
    rows = q.order_by(models.RutasHoy.orden).all()
    result = []
    for r in rows:
        cliente = r.cliente
        # filter out soft-deleted clientes from ruta view? keep but mark
        result.append(
            {
                "orden": r.orden,
                "cliente_id": r.cliente_id,
                "fecha": r.fecha.isoformat(),
                "entregado": bool(r.entregado),
                "delivered_at": r.delivered_at.isoformat() if r.delivered_at else None,
                "cliente": schemas.ClienteRead.model_validate(cliente).model_dump(mode="json") if cliente else None,
            }
        )
    return result


@app.patch("/rutas/hoy/entregado")
def marcar_entregado(payload: schemas.EntregadoRequest, db: Session = Depends(get_db)):
    if not payload.cliente_ids:
        raise HTTPException(status_code=400, detail="cliente_ids required")
    today = date.today()
    now = datetime.now(timezone.utc)
    rows = db.query(models.RutasHoy).filter(models.RutasHoy.fecha == today, models.RutasHoy.cliente_id.in_([str(x) for x in payload.cliente_ids])).all()
    updated = 0
    for r in rows:
        if not r.entregado:
            r.entregado = True
            r.delivered_at = now
            updated += 1
    db.commit()
    return {"updated": updated, "fecha": today.isoformat()}


@app.delete("/rutas/hoy")
def terminar_lista(db: Session = Depends(get_db)):
    today = date.today()
    pending = db.query(models.RutasHoy).filter(models.RutasHoy.fecha == today, models.RutasHoy.entregado == False).count()  # noqa: E712
    if pending > 0:
        raise HTTPException(status_code=409, detail=f"pending {pending} — entrega todos antes de terminar")
    deleted = db.query(models.RutasHoy).filter(models.RutasHoy.fecha == today).delete()
    db.commit()
    return {"deleted": deleted, "fecha": today.isoformat()}


@app.delete("/rutas/hoy/terminar")
def terminar_lista_alias(db: Session = Depends(get_db)):
    return terminar_lista(db)


# ---------------------------------------------------------------------------
# Import 20 -> 11 (openpyxl, NFD, RIF validation, dedup, flagged #)
# ---------------------------------------------------------------------------

def _cell_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s != "" else None


@app.post("/clientes/import")
def import_clientes(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > MAX_IMPORT_SIZE:
        raise HTTPException(status_code=413, detail="File too large — max 5MB")
    # Block clearly invalid content-types (e.g. text/image/json masquerading as xlsx)
    if file.content_type:
        ct = file.content_type.lower()
        if ct.startswith("text/") or ct.startswith("image/") or ct == "application/json":
            raise HTTPException(status_code=400, detail=f"Invalid content type: {ct}")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Empty workbook")
    if ws.max_row > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"Too many rows — max {MAX_IMPORT_ROWS}")

    existing = db.query(models.Cliente).filter(models.Cliente.deleted == 0).all()
    existing_rifs: set[str] = set()
    existing_normalized: set[str] = set()
    for c in existing:
        rif_val = (c.rif or c.rif_cedula or "")
        if rif_val:
            existing_rifs.add(rif_val.strip().upper())
        if c.nombre_normalizado:
            existing_normalized.add(c.nombre_normalizado)

    imported = 0
    skipped = 0
    duplicates = 0
    flagged = 0

    to_insert: list[models.Cliente] = []

    for r in range(2, ws.max_row + 1):
        nombre_raw = ws.cell(row=r, column=2).value  # B Nombre_Canonico
        if nombre_raw is None or str(nombre_raw).strip() == "":
            continue

        nombre_str = str(nombre_raw).strip()
        rif_raw = ws.cell(row=r, column=5).value  # E RIF
        rif_str = _cell_str(rif_raw)
        rif_upper: str | None = None
        if rif_str is not None:
            rif_upper = rif_str.upper()
            if not RIF_REGEX.match(rif_upper):
                skipped += 1
                continue
            rif_str = rif_upper

        normalized = normalize_nfd(nombre_str)

        empresa_raw = ws.cell(row=r, column=12).value  # L Empresa
        empresa_str = _cell_str(empresa_raw)
        is_flagged = nombre_str.startswith("#") or (empresa_str is not None and "#" in empresa_str)

        dup_by_rif = rif_upper is not None and rif_upper in existing_rifs
        dup_by_name = normalized in existing_normalized
        if dup_by_rif or dup_by_name:
            duplicates += 1
            continue

        direccion_raw = ws.cell(row=r, column=10).value  # J Direccion
        telefono_raw = ws.cell(row=r, column=8).value  # H Telefono
        zona_raw = ws.cell(row=r, column=13).value  # M Zona_Ruta

        direccion = _cell_str(direccion_raw)
        telefono = _cell_str(telefono_raw)
        zona = _cell_str(zona_raw)

        gen_id = generate_uuid()
        now_dt = datetime.now(timezone.utc)
        cliente = models.Cliente(
            id=gen_id,
            nombre=nombre_str,
            nombre_normalizado=normalized,
            rif=rif_str,
            rif_cedula=rif_str,
            telefono=telefono,
            direccion=direccion,
            direccion_original=direccion,  # frozen
            empresa=empresa_str,
            zona=zona,
            lat=None,
            lng=None,
            latitud=None,
            longitud=None,
            texto_breve=None,
            is_flagged=is_flagged,
            has_gps_fix=False,
            updated_at=now_dt,
            sync_status=0,
            deleted=0,
        )
        to_insert.append(cliente)
        if rif_upper:
            existing_rifs.add(rif_upper)
        existing_normalized.add(normalized)
        if is_flagged:
            flagged += 1

    if to_insert:
        db.add_all(to_insert)
        db.commit()
        imported = len(to_insert)

    return {"imported": imported, "skipped": skipped, "duplicates": duplicates, "flagged": flagged}


# ---------------------------------------------------------------------------
# Serve frontend static at /app (Leaflet)
# ---------------------------------------------------------------------------
_frontend_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend")
if os.path.isdir(_frontend_dir):
    app.mount("/app", StaticFiles(directory=_frontend_dir, html=True), name="frontend")

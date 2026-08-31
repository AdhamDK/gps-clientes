"""Tests for PR4 Optimize (VROOM MockTransport) + Health/CORS + 9 endpoints verify."""

import io
import time

import httpx
import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base, get_db
from backend.main import app


@pytest.fixture()
def client():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=engine)


def _create_clientes_with_gps(cli, n=5):
    ids = []
    for i in range(n):
        resp = cli.post(
            "/clientes",
            json={"nombre": f"Cliente {i} Zona", "lat": 8.61 + i * 0.01, "lng": -71.65 - i * 0.01, "zona": "VIGIA"},
        )
        assert resp.status_code == 201, resp.text
        ids.append(resp.json()["id"])
    return ids


def _create_cliente_no_gps(cli):
    resp = cli.post("/clientes", json={"nombre": "Sin GPS", "zona": "VIGIA"})
    assert resp.status_code == 201
    return resp.json()["id"]


# ---------------------------------------------------------------------------
# VROOM MockTransport helpers
# ---------------------------------------------------------------------------


def _vroom_success_handler(request: httpx.Request) -> httpx.Response:
    if request.url.path == "/" and request.method == "POST":
        import json as _json

        payload = _json.loads(request.content.decode() or "{}")
        jobs = payload.get("jobs", [])
        ordered = list(reversed([j["id"] for j in jobs])) if len(jobs) > 1 else [j["id"] for j in jobs]
        steps = [{"type": "start", "id": 0}]
        for jid in ordered:
            steps.append({"type": "job", "id": jid})
        steps.append({"type": "end", "id": 0})
        body = {
            "code": 0,
            "summary": {"distance": 12345, "duration": 3600},
            "routes": [{"vehicle": 0, "distance": 12345, "duration": 3600, "steps": steps}],
        }
        return httpx.Response(200, json=body)
    if "route" in request.url.path:
        geo = {"type": "LineString", "coordinates": [[-71.65, 8.61], [-71.66, 8.62]]}
        body = {"code": "Ok", "routes": [{"geometry": geo, "distance": 12345, "duration": 3600}]}
        return httpx.Response(200, json=body)
    if request.url.path in ("/health", "/"):
        return httpx.Response(200, json={"status": "ok"})
    return httpx.Response(404, json={"error": "not found"})


def _vroom_osrm_transport_success():
    return httpx.MockTransport(_vroom_success_handler)


# ---------------------------------------------------------------------------
# Optimize success / 400 / 502
# ---------------------------------------------------------------------------


def test_optimize_success_5_jobs_uses_mocktransport(client, monkeypatch):
    ids = _create_clientes_with_gps(client, 5)
    transport = _vroom_osrm_transport_success()

    import backend.vroom_client as vc

    orig_client = httpx.Client

    def _mock_client(*args, **kwargs):
        kwargs["transport"] = transport
        return orig_client(*args, **kwargs)

    monkeypatch.setattr(vc.httpx, "Client", _mock_client)

    resp = client.post("/rutas/optimizar", json={"cliente_ids": ids})
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert "orden" in data and len(data["orden"]) == 5
    assert data["distance"] > 0
    assert data["geometry"] is not None
    resp2 = client.get("/rutas/hoy")
    assert resp2.status_code == 200
    rows = resp2.json()
    assert len(rows) == 5
    assert [r["orden"] for r in rows] == list(range(5))


def test_optimize_hasGpsFix_false_400_no_vroom_call(client, monkeypatch):
    ids = _create_clientes_with_gps(client, 2)
    no_gps = _create_cliente_no_gps(client)
    ids_with_bad = ids[:1] + [no_gps]

    def _fail_handler(request: httpx.Request) -> httpx.Response:
        raise AssertionError("VROOM should not be called when hasGpsFix false")

    transport = httpx.MockTransport(_fail_handler)
    import backend.vroom_client as vc

    orig_client = httpx.Client

    def _mock_client(*args, **kwargs):
        kwargs["transport"] = transport
        return orig_client(*args, **kwargs)

    monkeypatch.setattr(vc.httpx, "Client", _mock_client)

    resp = client.post("/rutas/optimizar", json={"cliente_ids": ids_with_bad})
    assert resp.status_code == 400, resp.text
    assert "GPS" in resp.text or "has" in resp.text.lower()
    resp2 = client.get("/rutas/hoy")
    assert resp2.json() == []


def test_optimize_vroom_down_502_after_retry(client, monkeypatch):
    ids = _create_clientes_with_gps(client, 3)

    def _down_handler(request: httpx.Request) -> httpx.Response:
        raise httpx.ConnectError("VROOM down")

    transport = httpx.MockTransport(_down_handler)
    import backend.vroom_client as vc

    orig_client = httpx.Client

    def _mock_client(*args, **kwargs):
        kwargs["transport"] = transport
        return orig_client(*args, **kwargs)

    monkeypatch.setattr(vc.httpx, "Client", _mock_client)

    resp = client.post("/rutas/optimizar", json={"cliente_ids": ids})
    assert resp.status_code == 502, resp.text


def test_optimize_vroom_http_500_maps_to_502(client, monkeypatch):
    ids = _create_clientes_with_gps(client, 2)

    def _500_handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == "/" and request.method == "POST":
            return httpx.Response(500, json={"error": "internal"})
        return httpx.Response(200, json={"code": "Ok", "routes": [{"geometry": None}]})

    transport = httpx.MockTransport(_500_handler)
    import backend.vroom_client as vc

    orig_client = httpx.Client

    def _mock_client(*args, **kwargs):
        kwargs["transport"] = transport
        return orig_client(*args, **kwargs)

    monkeypatch.setattr(vc.httpx, "Client", _mock_client)

    resp = client.post("/rutas/optimizar", json={"cliente_ids": ids})
    assert resp.status_code == 502


def test_get_rutas_hoy_with_fecha_param(client):
    resp = client.get("/rutas/hoy")
    assert resp.status_code == 200
    assert resp.json() == []
    resp2 = client.get("/rutas/hoy?fecha=not-a-date")
    assert resp2.status_code == 400


# ---------------------------------------------------------------------------
# Health / CORS
# ---------------------------------------------------------------------------


def test_health_always_200_even_when_vroom_down(client, monkeypatch):
    import backend.vroom_client as vc

    monkeypatch.setattr(vc, "check_vroom", lambda transport=None, timeout=2.0: "down")
    monkeypatch.setattr(vc, "check_osrm", lambda transport=None, timeout=2.0: "down")

    resp = client.get("/health")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "ok"
    assert data["vroom"] == "down"
    assert data["osrm"] == "down"


def test_health_vroom_up(client, monkeypatch):
    import backend.vroom_client as vc

    monkeypatch.setattr(vc, "check_vroom", lambda transport=None, timeout=2.0: "up")
    monkeypatch.setattr(vc, "check_osrm", lambda transport=None, timeout=2.0: "up")

    resp = client.get("/health")
    assert resp.status_code == 200
    assert resp.json()["vroom"] == "up"
    assert resp.json()["osrm"] == "up"


def test_cors_allows_localhost_and_file_origin(client):
    resp = client.get("/health", headers={"Origin": "http://localhost:3000"})
    assert resp.status_code == 200
    assert "access-control-allow-origin" in {k.lower(): v for k, v in resp.headers.items()}
    resp2 = client.get("/health", headers={"Origin": "file://"})
    assert resp2.status_code == 200
    resp3 = client.options("/health", headers={"Origin": "http://localhost:5173", "Access-Control-Request-Method": "GET"})
    assert resp3.status_code in (200, 204)


# ---------------------------------------------------------------------------
# Verify 9 endpoints <2s per 10 clients (spec 6.2)
# ---------------------------------------------------------------------------


def test_verify_9_endpoints_under_2s_per_10_clients(client):
    start = time.perf_counter()
    assert client.get("/health").status_code == 200
    r = client.post("/clientes", json={"nombre": "Bench Uno", "lat": 8.61, "lng": -71.65})
    assert r.status_code == 201
    cid = r.json()["id"]
    assert client.get(f"/clientes/{cid}").status_code == 200
    assert client.get("/clientes?q=bench").status_code == 200
    assert client.patch(f"/clientes/{cid}", json={"zona": "BENCH"}).status_code == 200
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([""] * 13)
    ws.cell(row=1, column=2, value="Nombre")
    ws.cell(row=1, column=5, value="RIF")
    ws.cell(row=1, column=12, value="Empresa")
    ws.append(["", "Bench Import", "", "", "J12345678", "", "", "0414", "", "Dir", "", "EMP", "ZONA"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp_imp = client.post("/clientes/import", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp_imp.status_code == 200
    assert client.get("/clientes/export?formato=xlsx").status_code == 200
    assert client.get("/clientes/export?formato=pdf").status_code == 200
    assert client.get("/rutas/hoy").status_code == 200
    elapsed = time.perf_counter() - start
    assert elapsed < 2.0, f"9 endpoints took {elapsed:.3f}s >=2s"
    assert client.delete(f"/clientes/{cid}").status_code == 204


def test_vroom_client_direct_with_mocktransport():
    import backend.vroom_client as vc

    transport = _vroom_osrm_transport_success()
    jobs = [{"id": 1, "location": [-71.65, 8.61]}, {"id": 2, "location": [-71.66, 8.62]}]
    data = vc.optimize_via_vroom(jobs, transport=transport)
    assert data["code"] == 0
    assert len(data["routes"][0]["steps"]) == 4
    geo = vc.fetch_geometry([[-71.65, 8.61], [-71.66, 8.62]], transport=transport)
    assert geo["geometry"] is not None
    assert geo["distance"] > 0

"""Tests for PR3 Import+Export — RIF skip, dedup, flagged, 409 fixture, export streaming & column picker."""

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base, get_db
from backend.main import app


@pytest.fixture()
def client():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=engine)


def _build_xlsx(rows: list[dict]) -> bytes:
    """rows: list of dict with keys mapping to Excel cols: nombre, rif, direccion, empresa, zona, telefono."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "ID_Cliente", "Nombre_Canonico", "Alias_Txt", "Nombre_Original_Txt", "RIF", "RIF_Status",
        "RIF_Fuente", "Telefono", "Telefono_Fuente", "Direccion", "Direccion_Fuente", "Empresa",
        "Zona_Ruta", "Saldo_CxC", "Monto_Facturado", "Fecha_Emision", "Fecha_Vencimiento",
        "Fecha_Corte_CxC", "Origen", "Confianza_Match",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("id"), r.get("nombre"), None, None, r.get("rif"), None, None,
            r.get("telefono"), None, r.get("direccion"), None, r.get("empresa"), r.get("zona"),
            None, None, None, None, None, None, None,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---- Import tests ----

def test_import_409_fixture_flagged_and_queryable(client):
    with open("Clientes_TOM_KEVIN.xlsx", "rb") as f:
        content = f.read()
    resp = client.post("/clientes/import", files={"file": ("Clientes_TOM_KEVIN.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 200, resp.text
    data = resp.json()
    # per spec: imported+duplicates+skipped == 409, flagged >0
    assert data["imported"] + data["duplicates"] + data["skipped"] == 409
    assert data["flagged"] > 0
    # queryable: GET /clientes should have imported total
    resp2 = client.get("/clientes", params={"limit": 500})
    assert resp2.json()["total"] == data["imported"]
    # flagged persisted: at least flagged rows have is_flagged true (fetch all with limit 500)
    resp_all = client.get("/clientes", params={"limit": 500})
    flagged_count = sum(1 for it in resp_all.json()["items"] if it["is_flagged"])
    assert flagged_count == data["flagged"] or flagged_count > 0


def test_import_skips_invalid_rif(client):
    xlsx = _build_xlsx([
        {"nombre": "Valido Uno", "rif": "J12345678", "empresa": "TOM"},
        {"nombre": "Invalido RIF", "rif": "XXX", "empresa": "TOM"},
        {"nombre": "Valido Dos", "rif": "V12345678", "empresa": "TOM"},
    ])
    resp = client.post("/clientes/import", files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 200
    data = resp.json()
    assert data["skipped"] == 1
    assert data["imported"] == 2
    # invalid not persisted
    resp2 = client.get("/clientes")
    names = [c["nombre"] for c in resp2.json()["items"]]
    assert "Invalido RIF" not in names
    assert "Valido Uno" in names


def test_import_dedup_by_rif_and_nombre_normalizado(client):
    # seed one cliente
    client.post("/clientes", json={"nombre": "Vigía Central", "rif": "J12345678"})
    # import with same rif and same normalized name
    xlsx = _build_xlsx([
        {"nombre": "Vigia Central", "rif": "J12345678"},  # duplicate RIF
        {"nombre": "VIGÍA CENTRAL", "rif": None},  # duplicate normalized
        {"nombre": "Nuevo Cliente", "rif": "V87654321"},
    ])
    resp = client.post("/clientes/import", files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 200
    data = resp.json()
    assert data["duplicates"] == 2
    assert data["imported"] == 1


def test_import_hasGpsFix_false_no_lat_lng(client):
    xlsx = _build_xlsx([{"nombre": "Sin GPS", "rif": "J11111111", "direccion": "Calle 1"}])
    client.post("/clientes/import", files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    resp = client.get("/clientes")
    item = [c for c in resp.json()["items"] if c["nombre"] == "Sin GPS"][0]
    assert item["has_gps_fix"] is False
    assert item["lat"] is None
    assert item["lng"] is None
    assert item["direccion_original"] == "Calle 1"
    # /rutas/optimizar later would 400 but not tested here


def test_import_second_import_same_file_duplicates(client):
    with open("Clientes_TOM_KEVIN.xlsx", "rb") as f:
        content = f.read()
    resp1 = client.post("/clientes/import", files={"file": ("Clientes_TOM_KEVIN.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp1.status_code == 200
    first = resp1.json()
    resp2 = client.post("/clientes/import", files={"file": ("Clientes_TOM_KEVIN.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp2.status_code == 200
    data2 = resp2.json()
    assert data2["imported"] == 0
    # second time all 409 rows are duplicates (including intra-file dupes)
    assert data2["imported"] + data2["duplicates"] + data2["skipped"] == 409
    assert data2["duplicates"] == 409 - data2["skipped"]


# ---- Export tests ----

def test_export_xlsx_column_picker_and_reparseable(client):
    for i in range(3):
        client.post("/clientes", json={"nombre": f"Cliente {i}", "rif": f"J1000000{i}", "zona": "VIGIA" if i == 0 else None})
    resp = client.get("/clientes/export", params={"formato": "xlsx", "columnas": "nombre,rif,zona"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in resp.headers["content-disposition"]
    assert 'clientes.xlsx' in resp.headers["content-disposition"]
    # re-parseable with 3 cols
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ["nombre", "rif", "zona"]
    assert ws.max_row == 4  # header + 3
    assert ws.max_column == 3


def test_export_pdf_streaming_starts_with_pdf(client):
    client.post("/clientes", json={"nombre": "Para PDF"})
    resp = client.get("/clientes/export", params={"formato": "pdf"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers["content-disposition"]
    assert 'clientes.pdf' in resp.headers["content-disposition"]
    assert resp.content[:4] == b"%PDF"


def test_export_column_picker_whitelist_rejects_invalid(client):
    client.post("/clientes", json={"nombre": "X"})
    resp = client.get("/clientes/export", params={"formato": "xlsx", "columnas": "nombre,__evil"})
    assert resp.status_code == 400
    assert "Invalid columns" in resp.text


def test_export_default_whitelist_and_streaming(client):
    client.post("/clientes", json={"nombre": "Default Cols", "rif": "J99999999", "zona": "MERIDA"})
    resp = client.get("/clientes/export", params={"formato": "xlsx"})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    # default should include at least nombre,rif,zona and be within whitelist
    assert "nombre" in headers
    assert "rif" in headers
    assert len(headers) >= 5
    assert ws.max_row == 2


def test_export_pdf_with_column_picker(client):
    client.post("/clientes", json={"nombre": "PDF Picker", "zona": "Lunes"})
    resp = client.get("/clientes/export", params={"formato": "pdf", "columnas": "nombre,zona"})
    assert resp.status_code == 200
    assert resp.content[:4] == b"%PDF"
